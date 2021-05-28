import win32com.client as win32
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 打开工作表
wb = load_workbook('./material/04_月考勤表.xlsx')
sheet = wb.active

# 编写正文内容
content = '四月的考勤表已出，其中迟到时长超出 45 分钟的人员如下：\n'
for row_data in sheet.iter_rows(min_row=2, values_only=True):
    # 获取迟到时长超过45分钟的人员
    if row_data[2] > 45:
        content += '姓名：{name} 迟到总时长：{time} \n'.format(name=row_data[1], time=row_data[2])
content += '详情见附件内容'

 
def send_mail():
    outlook = win32.Dispatch('Outlook.Application')
 
    mail_item = outlook.CreateItem(0) # 0: create mail
 
    mail_item.Recipients.Add('del_fyodis@live.com')
    mail_item.Subject = '04_月考勤表'
 
    # mail_item.BodyFormat = 2          # 2: Html format
    # mail_item.HTMLBody  = '''
        # <H2>Hello, This is a test mail.</H2>
        # Hello Guys.
        # '''
        
    mail_item.Body = content
    mail_item.Attachments.Add('C:/Users/Del_F/Desktop/SoSe_2021/Pre-doctoral_Self_Study/Python/辦公相關/material/04_月考勤表.xlsx')   
    mail_item.Send()
 
if __name__ == '__main__':
    send_mail()
