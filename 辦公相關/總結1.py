from openpyxl import load_workbook

# 打开【practice_final.xlsx】工作簿
staff_wb = load_workbook('./material/practice_final.xlsx')
# 获取【上半年公司名单】工作表
staff_ws = staff_wb['上半年公司名单']

# 使用iter_rows()获取指定位置单元格的值，并打印
for row in staff_ws.iter_rows(min_row=11, min_col=2, max_col=3,max_row=13,values_only=True):
    print(row)