"""题目要求
1. 打开material文件夹下的文件practice1.xlsx，获取下半年公司名单工作表。
2. 然后打印出第5到第10行，前三列的所有数据（值）。
3. 再给工作表的最后一行添加数据'S1911', '萧爵瑟', 3000, '内容'。
4. 然后将结果保存在material文件夹下，并命名为`practice1_result.xlsx`。
"""
from openpyxl import load_workbook

# 打开【practice1.xlsx】工作簿
staff_wb = load_workbook('./material/practice1.xlsx')
# 按表名取表
staff_ws = staff_wb['下半年公司名单']

# 打印出第5到第10行，前三列的所有数据（值）
for i in staff_ws.iter_rows(min_row=5, max_row=10, max_col=3, values_only=True):
    print(i)

# 在工作表最后增加数据
info_tuple = ('S1911', '萧爵瑟', 3000, '内容')
staff_ws.append(info_tuple)

# 保存结果为【practice1_result.xlsx】
staff_wb.save('./material/practice1_result.xlsx')
