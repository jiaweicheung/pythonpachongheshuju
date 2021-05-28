"""题目要求
1. 打开material文件夹下的'practice0.xlsx'。
2. 打印获得的工作簿对象。
3. 新建一个工作簿。
4. 打印新创建的工作簿对象。
5. 将新工作簿保存在material文件夹下，命名为'practice0_result.xlsx'。 
"""
from openpyxl import load_workbook, Workbook

# 打开【practice0.xlsx】工作簿
practice_wb = load_workbook('./material/practice0.xlsx')
# 打印工作簿对象
print(practice_wb)

# 新建工作簿对象
new_wb = Workbook()
# 打印新建的工作簿对象
print(new_wb)

# 将新建的工作簿保存为【practice0_result.xlsx】
new_wb.save('./material/practice0_result.xlsx')